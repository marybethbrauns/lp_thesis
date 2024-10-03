import numpy as np
import pandas as pd
from tqdm import tqdm
import pulp
from deap import base, creator, tools, algorithms
import random
from sklearn.model_selection import train_test_split
import time
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

class IntegerLPProblem:
    def __init__(self, num_variables, num_constraints):
        self.num_variables = num_variables
        self.num_constraints = num_constraints
        self.objective = None
        self.constraints = None
        self.bounds = None
        self.solution = None

    def generate_problem(self):
        self.objective = np.random.randint(-10, 11, size=self.num_variables)
        self.constraints = np.random.randint(-5, 6, size=(self.num_constraints, self.num_variables))
        rhs = np.random.randint(1, 51, size=self.num_constraints)
        self.constraints = np.column_stack((self.constraints, rhs))
        self.bounds = [(0, 10) for _ in range(self.num_variables)]

    def solve(self):
        start_time = time.time()
        prob = pulp.LpProblem("Integer LP Problem", pulp.LpMinimize)
        vars = [pulp.LpVariable(f'x{i}', lowBound=self.bounds[i][0], upBound=self.bounds[i][1], cat='Integer')
                for i in range(self.num_variables)]
        prob += pulp.lpSum(self.objective[i] * vars[i] for i in range(self.num_variables))
        for i in range(self.num_constraints):
            prob += pulp.lpSum(self.constraints[i, j] * vars[j] for j in range(self.num_variables)) <= self.constraints[i, -1]
        prob.solve()
        end_time = time.time()
        self.solution = {
            'status': pulp.LpStatus[prob.status],
            'objective_value': pulp.value(prob.objective),
            'variables': [var.varValue for var in prob.variables()],
            'solve_time': end_time - start_time
        }

    def to_dict(self):
        return {
            'num_variables': self.num_variables,
            'num_constraints': self.num_constraints,
            'objective': self.objective.tolist(),
            'constraints': self.constraints.tolist(),
            'bounds': self.bounds,
            'solution': self.solution
        }

class LPDatasetGenerator:
    def __init__(self, num_problems, min_variables, max_variables, min_constraints, max_constraints):
        self.num_problems = num_problems
        self.min_variables = min_variables
        self.max_variables = max_variables
        self.min_constraints = min_constraints
        self.max_constraints = max_constraints
        self.problems = []

    def generate_dataset(self):
        for _ in tqdm(range(self.num_problems), desc="Generating problems"):
            num_vars = np.random.randint(self.min_variables, self.max_variables + 1)
            num_cons = np.random.randint(self.min_constraints, self.max_constraints + 1)
            problem = IntegerLPProblem(num_vars, num_cons)
            problem.generate_problem()
            problem.solve()
            self.problems.append(problem.to_dict())

    def to_dataframe(self):
        return pd.DataFrame(self.problems)

class GeneticAlgorithmSolver:
    def __init__(self, pop_size=50, ngen=50, cxpb=0.7, mutpb=0.2):
        self.pop_size = pop_size
        self.ngen = ngen
        self.cxpb = cxpb
        self.mutpb = mutpb

    @staticmethod
    def evaluate(individual, problem):
        objective = problem['objective']
        constraints = problem['constraints']

        obj_value = sum(x * c for x, c in zip(individual, objective))

        penalty = 0
        for constraint in constraints:
            left_side = sum(x * c for x, c in zip(individual, constraint[:-1]))
            if left_side > constraint[-1]:
                penalty += (left_side - constraint[-1]) * 100

        return obj_value + penalty,

    @staticmethod
    def calculate_objective_value(solution, problem):
        return sum(x * c for x, c in zip(solution, problem['objective']))

    def create_toolbox(self, num_variables, bounds):
        creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
        creator.create("Individual", list, fitness=creator.FitnessMin)

        toolbox = base.Toolbox()
        toolbox.register("attr_int", random.randint, bounds[0], bounds[1])
        toolbox.register("individual", tools.initRepeat, creator.Individual, toolbox.attr_int, n=num_variables)
        toolbox.register("population", tools.initRepeat, list, toolbox.individual)

        toolbox.register("evaluate", self.evaluate)
        toolbox.register("mate", tools.cxTwoPoint)
        toolbox.register("mutate", tools.mutUniformInt, low=bounds[0], up=bounds[1], indpb=0.1)
        toolbox.register("select", tools.selTournament, tournsize=3)

        return toolbox

    def solve(self, problem):
        num_variables = problem['num_variables']
        bounds = problem['bounds'][0]  # Assuming all variables have the same bounds

        toolbox = self.create_toolbox(num_variables, bounds)
        toolbox.register("evaluate", self.evaluate, problem=problem)

        pop = toolbox.population(n=self.pop_size)
        hof = tools.HallOfFame(1)
        stats = tools.Statistics(lambda ind: ind.fitness.values)
        stats.register("avg", np.mean)
        stats.register("min", np.min)
        stats.register("max", np.max)

        start_time = time.time()
        pop, log = algorithms.eaSimple(pop, toolbox, cxpb=self.cxpb, mutpb=self.mutpb,
                                       ngen=self.ngen, stats=stats, halloffame=hof, verbose=False)
        end_time = time.time()

        return {
            'best_solution': hof[0],
            'best_fitness': hof[0].fitness.values[0],
            'predicted_objective': self.calculate_objective_value(hof[0], problem),
            'solve_time': end_time - start_time
        }

class Experiment:
    def __init__(self, num_problems, min_variables, max_variables, min_constraints, max_constraints):
        self.dataset_generator = LPDatasetGenerator(num_problems, min_variables, max_variables,
                                                    min_constraints, max_constraints)
        self.ga_solver = GeneticAlgorithmSolver()
        self.problems_df = None
        self.train_df = None
        self.test_df = None
        self.train_results = None
        self.test_results = None
        self.results_combined = None

    def generate_dataset(self):
        self.dataset_generator.generate_dataset()
        self.problems_df = self.dataset_generator.to_dataframe()

    def split_dataset(self, test_size=0.2, random_state=42):
        self.train_df, self.test_df = train_test_split(self.problems_df, test_size=test_size, random_state=random_state)

    def process_dataset(self, df, is_training=True):
        results = []
        for _, problem in tqdm(df.iterrows(), total=len(df), desc="Processing"):
            ga_result = self.ga_solver.solve(problem)
            results.append({
                'problem_id': _,
                'best_solution': ga_result['best_solution'],
                'best_fitness': ga_result['best_fitness'],
                'predicted_objective': ga_result['predicted_objective'],
                'ga_solve_time': ga_result['solve_time'],
                'optimal_solution': problem['solution']['variables'],
                'optimal_objective': problem['solution']['objective_value'],
                'pulp_solve_time': problem['solution']['solve_time']
            })
        return pd.DataFrame(results)

    def run_experiment(self):
        self.generate_dataset()
        self.split_dataset()
        print(f"Training set size: {len(self.train_df)}")
        print(f"Testing set size: {len(self.test_df)}")
        self.train_results = self.process_dataset(self.train_df, is_training=True)
        self.test_results = self.process_dataset(self.test_df, is_training=False)

    def calculate_metrics(self):
        train_mae = np.mean(np.abs(self.train_results['predicted_objective'] - self.train_results['optimal_objective']))
        test_mae = np.mean(np.abs(self.test_results['predicted_objective'] - self.test_results['optimal_objective']))
        print(f"\nTraining Mean Absolute Error: {train_mae}")
        print(f"Testing Mean Absolute Error: {test_mae}")

        self.results_combined = pd.DataFrame({
            'problem_id': self.test_results['problem_id'],
            'predicted_objective_value': self.test_results['predicted_objective'],
            'actual_objective_value': self.test_results['optimal_objective'],
            'solution_fitness': self.test_results['best_fitness'],
            'ga_solve_time': self.test_results['ga_solve_time'],
            'pulp_solve_time': self.test_results['pulp_solve_time']
        })

        self.results_combined['time_difference'] = self.results_combined['pulp_solve_time'] - self.results_combined['ga_solve_time']
        self.results_combined['ga_faster'] = self.results_combined['time_difference'] > 0
        self.results_combined['speedup_factor'] = self.results_combined['pulp_solve_time'] / self.results_combined['ga_solve_time']

        mismatches = self.results_combined[self.results_combined['predicted_objective_value'] != self.results_combined['actual_objective_value']]
        total_problems = len(self.results_combined)
        mismatch_percentage = (len(mismatches) / total_problems) * 100
        overall_avg_difference = np.mean(np.abs(self.results_combined['predicted_objective_value'] - self.results_combined['actual_objective_value']))
        ga_faster_count = self.results_combined['ga_faster'].sum()
        avg_speedup = self.results_combined['speedup_factor'].mean()
        median_speedup = self.results_combined['speedup_factor'].median()

        return {
            'train_mae': train_mae,
            'test_mae': test_mae,
            'mismatch_percentage': mismatch_percentage,
            'overall_avg_difference': overall_avg_difference,
            'ga_faster_count': ga_faster_count,
            'ga_faster_percentage': (ga_faster_count / total_problems) * 100,
            'avg_speedup': avg_speedup,
            'median_speedup': median_speedup,
            'mismatches': mismatches,
            'total_problems': total_problems
        }

    def generate_excel_report(self, metrics):
        wb = Workbook()

        # Explanation Sheet
        ws_explanation = wb.active
        ws_explanation.title = "Explanation"
        self._add_explanation(ws_explanation)

        # Master List of Problems
        ws_master = wb.create_sheet("Master List of Problems")
        self._add_master_list(ws_master)

        # Overall Metrics Sheet
        ws_metrics = wb.create_sheet("Overall Metrics")
        self._add_overall_metrics(ws_metrics, metrics)

        # Combined Results Sheet
        ws_combined = wb.create_sheet("Combined Results")
        self._add_combined_results(ws_combined)

        # Top 5 Speedup Sheet
        ws_top_speedup = wb.create_sheet("Top 5 Speedup")
        self._add_top_speedup(ws_top_speedup)

        # Bottom 5 Speedup Sheet
        ws_bottom_speedup = wb.create_sheet("Bottom 5 Speedup")
        self._add_bottom_speedup(ws_bottom_speedup)

        # Mismatches Sheet
        ws_mismatches = wb.create_sheet("Mismatches")
        self._add_mismatches(ws_mismatches, metrics)

        # Apply styling to all sheets
        for sheet in wb:
            self._style_sheet(sheet)

        # Save the workbook
        excel_filename = '/Users/marybeth/Desktop/p_solver_results.xlsx'
        os.makedirs(os.path.dirname(excel_filename), exist_ok=True)
        wb.save(excel_filename)
        print(f"\nKey metrics and results have been saved to {excel_filename}")

    def _add_explanation(self, ws):
        explanations = [
            ("Sheet", "Description"),
            ("Explanation", "This sheet, providing an overview of the workbook contents."),
            ("Master List of Problems",
             "A comprehensive list of all problems generated, including their objectives, constraints, and solutions."),
            ("Overall Metrics", "Key performance metrics of the Genetic Algorithm (GA) compared to PuLP."),
            ("Combined Results",
             "Detailed results for each problem in the test set, comparing GA and PuLP performance."),
            ("Top 5 Speedup", "The 5 problems where GA had the highest speedup compared to PuLP."),
            ("Bottom 5 Speedup", "The 5 problems where GA had the lowest speedup compared to PuLP."),
            ("Mismatches", "Details of problems where GA's solution didn't match PuLP's optimal solution."),
            ("", ""),
            ("Key Terms", "Explanation"),
            ("Objective", "The function to be minimized in the linear programming problem."),
            ("Constraints", "The conditions that the solution must satisfy."),
            ("GA", "Genetic Algorithm, a heuristic method used to solve optimization problems."),
            ("PuLP", "An exact solver used as a benchmark for the GA's performance."),
            ("Speedup Factor", "The ratio of PuLP solve time to GA solve time. A factor > 1 means GA was faster."),
            ("MAE",
             "Mean Absolute Error, measuring the average difference between predicted and actual objective values.")
        ]
        for row in explanations:
            ws.append(row)

    def _add_master_list(self, ws):
        headers = ["Problem ID", "Num Variables", "Num Constraints", "Objective", "Constraints", "Bounds",
                   "Optimal Solution", "Optimal Objective Value"]
        ws.append(headers)
        for _, problem in self.problems_df.iterrows():
            row = [
                _,
                problem['num_variables'],
                problem['num_constraints'],
                str(problem['objective']),
                str(problem['constraints']),
                str(problem['bounds']),
                str(problem['solution']['variables']),
                problem['solution']['objective_value']
            ]
            ws.append(row)

    def _add_overall_metrics(self, ws, metrics):
        metrics_data = [
            ("Metric", "Value"),
            ("Training Mean Absolute Error", metrics['train_mae']),
            ("Testing Mean Absolute Error", metrics['test_mae']),
            ("Percentage of problems where GA didn't find the optimal solution", metrics['mismatch_percentage']),
            ("Overall average absolute difference", metrics['overall_avg_difference']),
            ("Number of problems where GA was faster", metrics['ga_faster_count']),
            ("Percentage of problems where GA was faster", metrics['ga_faster_percentage']),
            ("Average speedup factor (PuLP time / GA time)", metrics['avg_speedup']),
            ("Median speedup factor", metrics['median_speedup'])
        ]
        for row in metrics_data:
            ws.append(row)

    def _add_combined_results(self, ws):
        for r in dataframe_to_rows(self.results_combined, index=False, header=True):
            ws.append(r)

    def _add_top_speedup(self, ws):
        top_speedup = self.results_combined.nlargest(5, 'speedup_factor')[
            ['problem_id', 'speedup_factor', 'ga_solve_time', 'pulp_solve_time']]
        for r in dataframe_to_rows(top_speedup, index=False, header=True):
            ws.append(r)

    def _add_bottom_speedup(self, ws):
        bottom_speedup = self.results_combined.nsmallest(5, 'speedup_factor')[
            ['problem_id', 'speedup_factor', 'ga_solve_time', 'pulp_solve_time']]
        for r in dataframe_to_rows(bottom_speedup, index=False, header=True):
            ws.append(r)

    def _add_mismatches(self, ws, metrics):
        if not metrics['mismatches'].empty:
            for r in dataframe_to_rows(metrics['mismatches'], index=False, header=True):
                ws.append(r)
        else:
            ws['A1'] = "No mismatches found. The genetic algorithm found the optimal solution for all test problems."

    def _style_sheet(self, ws):
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        table = Table(displayName=f"Table_{ws.title.replace(' ', '_')}", ref=ws.dimensions)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)


def main():
    # Set random seeds for reproducibility
    np.random.seed(42)
    random.seed(42)

    # Create and run the experiment
    experiment = Experiment(num_problems=1000, min_variables=2, max_variables=5,
                            min_constraints=2, max_constraints=5)

    print("Starting experiment...")
    experiment.run_experiment()

    print("Calculating metrics...")
    metrics = experiment.calculate_metrics()

    print("Generating Excel report...")
    experiment.generate_excel_report(metrics)

    print("Experiment completed. Check the Excel file for detailed results.")

    # Print some key metrics to console
    print("\nKey Metrics:")
    print(f"Training Mean Absolute Error: {metrics['train_mae']:.4f}")
    print(f"Testing Mean Absolute Error: {metrics['test_mae']:.4f}")
    print(f"Percentage of problems where GA didn't find the optimal solution: {metrics['mismatch_percentage']:.2f}%")
    print(f"Average speedup factor (PuLP time / GA time): {metrics['avg_speedup']:.2f}")
    print(f"Percentage of problems where GA was faster: {metrics['ga_faster_percentage']:.2f}%")


if __name__ == "__main__":
    main()